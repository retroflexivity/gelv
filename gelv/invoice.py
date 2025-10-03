import openpyxl as xl
from copy import copy
from datetime import date
from typing import Optional
from io import BytesIO
from gelv.models import Payment, AbstractOrder
from gelv.utils import verbalize_price, trace


class Invoice:
    """
    An XLSX invoice class.
    """
    payment: Payment

    def __init__(self, payment) -> None:
        self.payment = payment

    @property
    def number(self) -> str:
        return self.payment.number

    @property
    def filename(self) -> str:
        return f'{self.number}.xlsx'

    @staticmethod
    def _copy_cell(src: xl.cell.Cell, dest: xl.cell.Cell, value: Optional[str] = None) -> None:
        if value:
            dest.value = value
        else:
            dest.value = src.value

        if src.has_style:
            dest.font = copy(src.font)  # type: ignore
            dest.font = copy(src.font)  # type: ignore
            dest.border = copy(src.border)  # type: ignore
            dest.fill = copy(src.fill)  # type: ignore
            dest.number_format = src.number_format  # type: ignore
            dest.protection = copy(src.protection)  # type: ignore
            dest.alignment = copy(src.alignment)  # type: ignore

    def generate(self) -> BytesIO:
        product_ws = xl.open('gelv/static/gelv/xlsx/product_row.xlsx')['product']

        wb = xl.open('gelv/static/gelv/xlsx/invoice.xlsx')
        ws = wb['invoice']
        ws['B1'] = self.number
        ws['B2'] = format(date.today(), '%d.%m.%Y')

        ws['B15'] = self.payment.name
        ws['B16'] = self.payment.personal_code
        ws['B17'] = self.payment.address
        ws['B18'] = self.payment.billing_email
        ws['B19'] = self.payment.phone
        ws['D19'] = self.payment.billing_email  # TODO: find out whether this is right

        ws['E23'] = format(self.payment.total_price, '.2f')
        ws['B25'] = verbalize_price(self.payment.total_price)

        order: AbstractOrder
        for order in list(self.payment.issueorder_set.all()) + list(self.payment.subscriptionorder_set.all()):
            ws.insert_rows(22)
            for src_ix, ref_ix, value in zip(
                ('A1', 'B1', 'C1', 'D1', 'E1'),
                ('A22', 'B22', 'C22', 'D22', 'E22'),
                (order, '', amount := getattr(order, 'amount', 1), format(order.price, '.2f'), amount * order.price)
            ):
                self._copy_cell(product_ws[src_ix], ws[ref_ix], value=str(value))

        wb.save(buffer := BytesIO())
        buffer.seek(0)
        return buffer
